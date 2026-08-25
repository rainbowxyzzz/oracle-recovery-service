from __future__ import annotations

import csv
import hashlib
import json
import shutil
import time
import uuid
from datetime import timedelta
from pathlib import Path
from typing import Any, Iterator

from sqlalchemy import desc, select

from recovery_service.api.schemas.doris_sql_etl import QueryExportStatus
from recovery_service.common.time import app_now
from recovery_service.core.models.task import DatabaseConnectionProfile, QueryExportJob
from recovery_service.db.session import get_sync_session_factory
from recovery_service.services.auth import AuthContext
from recovery_service.services.doris_sql_etl import _clean_select_sql, _doris_conn, _ensure_doris_profile, _jsonable
from recovery_service.settings import get_settings

_FORMATS = {"csv", "tsv", "jsonl", "xlsx", "parquet"}
_ENCODINGS = {"utf-8", "utf-8-sig", "gbk"}


class ExportCancelled(Exception):
    pass


def create_query_export_job(
    *, profile: DatabaseConnectionProfile, database: str | None, sql: str, export_format: str,
    encoding: str, resource_profile: str, actor: AuthContext,
) -> QueryExportStatus:
    _ensure_doris_profile(profile)
    clean_sql = _clean_export_sql(sql)
    fmt = str(export_format).lower()
    if fmt not in _FORMATS:
        raise ValueError("不支持的导出格式。")
    if encoding not in _ENCODINGS:
        raise ValueError("不支持的文本编码。")
    if resource_profile not in {"streaming", "columnar"}:
        raise ValueError("不支持的资源档位。")
    if fmt == "parquet" and resource_profile != "columnar":
        resource_profile = "columnar"
    now = app_now()
    job = QueryExportJob(
        id=uuid.uuid4(), connection_id=profile.id, connection_name=profile.name,
        database=(database or "").strip() or None, sql_text=clean_sql,
        sql_summary=_sql_summary(clean_sql), export_format=fmt, encoding=encoding if fmt in {"csv", "tsv"} else None,
        resource_profile=resource_profile, state="queued", message="导出任务已进入队列。",
        created_by_user_id=uuid.UUID(actor.user_id) if actor.user_id else None,
        created_by_username=actor.username, created_by_auth_type=actor.auth_type,
        created_at=now, updated_at=now,
    )
    session = get_sync_session_factory()()
    try:
        session.add(job)
        session.commit()
        task_id = _enqueue_export_worker(job.id)
        job.celery_task_id = task_id
        session.commit()
        session.refresh(job)
        return _to_status(job)
    finally:
        session.close()


def list_query_export_jobs(*, actor: AuthContext, limit: int = 100) -> list[QueryExportStatus]:
    session = get_sync_session_factory()()
    try:
        stmt = select(QueryExportJob).order_by(desc(QueryExportJob.created_at)).limit(limit)
        if not actor.is_admin:
            if actor.user_id:
                stmt = stmt.where(QueryExportJob.created_by_user_id == uuid.UUID(actor.user_id))
            else:
                stmt = stmt.where(QueryExportJob.created_by_username == actor.username)
        return [_to_status(item) for item in session.execute(stmt).scalars().all()]
    finally:
        session.close()


def get_query_export_job(job_id: uuid.UUID, *, actor: AuthContext) -> QueryExportJob:
    session = get_sync_session_factory()()
    try:
        job = session.get(QueryExportJob, job_id)
        if not job:
            raise KeyError("导出任务不存在。")
        if not _can_access(job, actor):
            raise PermissionError("无权访问该导出任务。")
        session.expunge(job)
        return job
    finally:
        session.close()


def record_query_export_download(job_id: uuid.UUID, *, actor: AuthContext) -> QueryExportJob:
    session = get_sync_session_factory()()
    try:
        job = session.get(QueryExportJob, job_id)
        if not job:
            raise KeyError("导出任务不存在。")
        if not _can_access(job, actor):
            raise PermissionError("无权下载该导出文件。")
        if job.state != "succeeded" or not job.file_path:
            raise ValueError("导出文件尚未生成完成。")
        if job.expires_at and job.expires_at <= app_now():
            _expire_job_file(job)
            session.commit()
            raise ValueError("导出文件已过期。")
        path = _safe_export_path(job.file_path)
        if not path.is_file():
            raise ValueError("导出文件不存在或已被清理。")
        job.download_count += 1
        job.downloaded_by_username = actor.username
        job.updated_at = app_now()
        session.commit()
        session.refresh(job)
        session.expunge(job)
        return job
    finally:
        session.close()


def cancel_query_export_job(job_id: uuid.UUID, *, actor: AuthContext) -> QueryExportStatus:
    session = get_sync_session_factory()()
    try:
        job = session.get(QueryExportJob, job_id)
        if not job:
            raise KeyError("导出任务不存在。")
        if not _can_access(job, actor):
            raise PermissionError("无权取消该导出任务。")
        if job.state not in {"queued", "running"}:
            raise ValueError("仅可取消排队中或运行中的导出任务。")
        job.state, job.current_stage, job.message = "cancelled", "cancelled", "导出任务已取消。"
        job.finished_at, job.updated_at = app_now(), app_now()
        session.commit()
        session.refresh(job)
        return _to_status(job)
    finally:
        session.close()


def cleanup_expired_query_exports() -> int:
    session = get_sync_session_factory()()
    try:
        jobs = session.execute(select(QueryExportJob).where(QueryExportJob.state == "succeeded", QueryExportJob.expires_at <= app_now())).scalars().all()
        for job in jobs:
            _expire_job_file(job)
        if jobs:
            session.commit()
        return len(jobs)
    finally:
        session.close()


def run_query_export(job_id: str) -> dict[str, Any]:
    job_uuid = uuid.UUID(job_id)
    session = get_sync_session_factory()()
    try:
        job = session.get(QueryExportJob, job_uuid)
        if not job or job.state != "queued":
            return {"state": "ignored"}
        profile = session.get(DatabaseConnectionProfile, job.connection_id)
        if not profile:
            raise ValueError("导出数据连接不存在。")
        session.expunge(profile)
        job.state, job.current_stage, job.message, job.started_at, job.updated_at = "running", "reading", "正在流式读取并生成导出文件。", app_now(), app_now()
        session.commit()
    finally:
        session.close()
    try:
        cleanup_expired_query_exports()
        _check_export_capacity()
        rows, byte_size, digest, file_name, file_path = _write_export(profile, job_uuid)
        session = get_sync_session_factory()()
        try:
            job = session.get(QueryExportJob, job_uuid)
            if job:
                job.state, job.current_stage, job.message = "succeeded", "completed", f"导出完成：{rows} 行，{byte_size} 字节。"
                job.row_count, job.processed_rows, job.progress_percent, job.byte_size, job.sha256 = rows, rows, 100, byte_size, digest
                job.file_name, job.file_path = file_name, str(file_path)
                job.expires_at = app_now() + timedelta(hours=max(1, get_settings().query_export_expire_hours))
                job.finished_at, job.updated_at = app_now(), app_now()
                session.commit()
        finally:
            session.close()
        return {"state": "succeeded", "rows": rows}
    except ExportCancelled:
        return {"state": "cancelled"}
    except Exception as exc:
        session = get_sync_session_factory()()
        try:
            job = session.get(QueryExportJob, job_uuid)
            if job:
                job.state, job.current_stage, job.message, job.error_message = "failed", "failed", f"导出失败：{exc}", str(exc)
                job.finished_at, job.updated_at = app_now(), app_now()
                session.commit()
        finally:
            session.close()
        return {"state": "failed", "message": str(exc)}


def _write_export(profile: DatabaseConnectionProfile, job_id: uuid.UUID) -> tuple[int, int, str, str, Path]:
    session = get_sync_session_factory()()
    try:
        job = session.get(QueryExportJob, job_id)
        if not job:
            raise ValueError("导出任务不存在。")
        fmt, database, sql, encoding = job.export_format, job.database, job.sql_text, job.encoding or "utf-8"
    finally:
        session.close()
    directory = _export_dir()
    directory.mkdir(parents=True, exist_ok=True)
    suffix = {"csv": "csv", "tsv": "tsv", "jsonl": "jsonl", "xlsx": "xlsx", "parquet": "parquet"}[fmt]
    file_name = f"query-export-{job_id.hex[:12]}.{suffix}"
    target = directory / file_name
    partial = directory / f".{file_name}.part"
    if partial.exists():
        partial.unlink()
    try:
        with _doris_conn(profile, database) as db, db.cursor() as cur:
            cur.execute(sql)
            if not cur.description:
                raise ValueError("导出 SQL 必须返回结果集。")
            columns = [str(item[0]) for item in cur.description]
            rows = _iter_rows(cur)
            count = _write_rows(partial, fmt, encoding, columns, rows, lambda processed: _record_progress(job_id, partial, processed))
        partial.replace(target)
        return count, target.stat().st_size, _sha256_file(target), file_name, target
    except Exception:
        partial.unlink(missing_ok=True)
        raise


def _iter_rows(cur) -> Iterator[dict[str, Any]]:
    batch_size = max(100, min(20000, get_settings().query_export_batch_size))
    while True:
        rows = cur.fetchmany(batch_size)
        if not rows:
            return
        for row in rows:
            yield {str(key): _jsonable(value) for key, value in row.items()}


def _write_rows(path: Path, fmt: str, encoding: str, columns: list[str], rows: Iterator[dict[str, Any]], progress) -> int:
    if fmt in {"csv", "tsv"}:
        with path.open("w", encoding=encoding, newline="") as stream:
            writer = csv.DictWriter(stream, fieldnames=columns, delimiter="\t" if fmt == "tsv" else ",", extrasaction="ignore")
            writer.writeheader()
            return _write_delimited(writer, rows, progress)
    if fmt == "jsonl":
        with path.open("w", encoding="utf-8", newline="\n") as stream:
            count = 0
            for row in rows:
                stream.write(json.dumps(row, ensure_ascii=False, default=str) + "\n")
                count += 1
                if count % 1000 == 0:
                    progress(count)
            progress(count)
            return count
    if fmt == "xlsx":
        from openpyxl import Workbook
        workbook = Workbook(write_only=True)
        sheet = workbook.create_sheet("查询结果")
        sheet.append(columns)
        count = 0
        for row in rows:
            sheet.append([row.get(column) for column in columns])
            count += 1
            if count % 1000 == 0:
                progress(count)
        workbook.save(path)
        progress(count)
        return count
    if fmt == "parquet":
        import pyarrow as pa
        import pyarrow.parquet as pq
        writer = None
        count = 0
        batch: list[dict[str, Any]] = []
        try:
            for row in rows:
                batch.append(row)
                if len(batch) >= 5000:
                    table = pa.Table.from_pylist(batch)
                    writer = writer or pq.ParquetWriter(path, table.schema)
                    writer.write_table(table)
                    count += len(batch)
                    progress(count)
                    batch = []
            if batch:
                table = pa.Table.from_pylist(batch)
                writer = writer or pq.ParquetWriter(path, table.schema)
                writer.write_table(table)
                count += len(batch)
                progress(count)
            elif writer is None:
                pq.write_table(pa.Table.from_pydict({column: [] for column in columns}), path)
            return count
        finally:
            if writer:
                writer.close()
    raise ValueError("不支持的导出格式。")


def _write_delimited(writer, rows: Iterator[dict[str, Any]], progress) -> int:
    count = 0
    for row in rows:
        writer.writerow(row)
        count += 1
        if count % 1000 == 0:
            progress(count)
    progress(count)
    return count


def _record_progress(job_id: uuid.UUID, partial_path: Path, processed_rows: int) -> None:
    session = get_sync_session_factory()()
    try:
        job = session.get(QueryExportJob, job_id)
        if not job or job.state == "cancelled":
            raise ExportCancelled()
        started = job.started_at or app_now()
        elapsed = max(0.001, (app_now() - started).total_seconds())
        job.processed_rows = processed_rows
        job.byte_size = partial_path.stat().st_size if partial_path.exists() else 0
        job.current_stage = "writing"
        job.throughput_rows_per_second = round(processed_rows / elapsed, 2)
        job.message = f"正在导出：已处理 {processed_rows} 行，{job.byte_size} 字节。"
        job.updated_at = app_now()
        session.commit()
    finally:
        session.close()


def _clean_export_sql(sql: str) -> str:
    clean = _clean_select_sql(sql)
    lowered = clean.lower()
    if any(token in lowered for token in (";", " insert ", " update ", " delete ", " drop ", " alter ", " create ")):
        raise ValueError("查询导出仅允许单条只读 SELECT / WITH SQL。")
    return clean


def _sql_summary(sql: str) -> str:
    return " ".join(sql.split())[:512]


def _export_dir() -> Path:
    return Path(get_settings().query_export_dir).resolve()


def _safe_export_path(value: str) -> Path:
    path = Path(value).resolve()
    if _export_dir() not in path.parents:
        raise ValueError("非法导出文件路径。")
    return path


def _check_export_capacity() -> None:
    _export_dir().mkdir(parents=True, exist_ok=True)
    usage = shutil.disk_usage(_export_dir())
    if usage.free < 512 * 1024 * 1024:
        raise RuntimeError("导出目录可用磁盘空间低于 512MB。")
    budget = _memory_budget_bytes()
    current = _current_memory_bytes()
    if current and current > budget:
        raise RuntimeError("导出 Worker 内存已超过任务预算，拒绝继续执行。")


def _current_memory_bytes() -> int:
    try:
        with open("/sys/fs/cgroup/memory.current", encoding="utf-8") as stream:
            return int(stream.read().strip())
    except (OSError, ValueError):
        return 0


def _memory_budget_bytes() -> int:
    configured = max(128, get_settings().query_export_memory_budget_mb) * 1024 * 1024
    try:
        with open("/sys/fs/cgroup/memory.max", encoding="utf-8") as stream:
            limit = stream.read().strip()
        if limit != "max":
            return min(configured, max(128 * 1024 * 1024, int(limit) // 2))
    except (OSError, ValueError):
        pass
    return configured


def _sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _can_access(job: QueryExportJob, actor: AuthContext) -> bool:
    if actor.is_admin:
        return True
    return bool(actor.user_id and job.created_by_user_id and str(job.created_by_user_id) == actor.user_id) or (not actor.user_id and job.created_by_username == actor.username)


def _expire_job_file(job: QueryExportJob) -> None:
    if job.file_path:
        _safe_export_path(job.file_path).unlink(missing_ok=True)
    job.state, job.message, job.file_path = "expired", "导出文件已过期并清理。", None
    job.updated_at = app_now()


def _enqueue_export_worker(job_id: uuid.UUID) -> str | None:
    from recovery_service.workers.celery_app import celery_app
    settings = get_settings()
    return celery_app.send_task("query_export.run", args=[str(job_id)], queue=settings.celery_data_export_queue).id


def _to_status(job: QueryExportJob) -> QueryExportStatus:
    return QueryExportStatus(
        job_id=job.id, connection_name=job.connection_name, database=job.database, sql_summary=job.sql_summary,
        export_format=job.export_format, encoding=job.encoding, resource_profile=job.resource_profile,
        state=job.state, message=job.message, row_count=job.row_count, byte_size=job.byte_size,
        processed_rows=job.processed_rows, progress_percent=job.progress_percent, current_stage=job.current_stage,
        throughput_rows_per_second=job.throughput_rows_per_second,
        sha256=job.sha256, file_name=job.file_name, error_message=job.error_message,
        created_by_username=job.created_by_username, download_count=job.download_count, expires_at=job.expires_at,
        created_at=job.created_at, started_at=job.started_at, updated_at=job.updated_at, finished_at=job.finished_at,
    )
